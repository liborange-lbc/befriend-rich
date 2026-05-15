import logging
from dataclasses import dataclass, field
from datetime import date, timedelta

from sqlalchemy.orm import Session

from app.models.fund import Fund
from app.models.import_log import ImportLog
from app.models.portfolio import PortfolioRecord
from app.services.market_data.exchange_rate import get_latest_rate
from app.services.portfolio.snapshot import generate_snapshot
from app.services.webank.classifier import classify_funds_with_ai
from app.services.webank.fund_matcher import find_or_create_fund

logger = logging.getLogger(__name__)


class DuplicateImportError(Exception):
    """Raised when a duplicate import is detected and force is not set."""

    pass


@dataclass(frozen=True)
class ImportResult:
    total_items: int
    matched_funds: int
    new_funds_created: int
    records_imported: int
    classification_results: dict = field(default_factory=dict)
    snapshot_generated: bool = False
    import_log_id: int = 0


def normalize_to_monday(d: date) -> date:
    """Normalize a date to the Monday of its ISO week."""
    return d - timedelta(days=d.weekday())


def _parse_excel(file_content: bytes) -> list[dict]:
    """
    Parse WeBank Excel file and extract asset items.

    Returns:
        List of {"资产项": str, "金额(元)": float, "币种": str}
    """
    from openpyxl import load_workbook
    from io import BytesIO

    wb = load_workbook(BytesIO(file_content), read_only=True)

    # Try to find "资产概览" sheet, fallback to first sheet
    if "资产概览" in wb.sheetnames:
        ws = wb["资产概览"]
    else:
        ws = wb.active

    rows_data = list(ws.iter_rows(values_only=True))
    wb.close()

    if not rows_data:
        raise ValueError("Excel 文件为空")

    # Find header row
    header = None
    header_idx = -1
    for i, row in enumerate(rows_data):
        row_strs = [str(cell).strip() if cell else "" for cell in row]
        if "资产项" in row_strs and any("金额" in s for s in row_strs):
            header = row_strs
            header_idx = i
            break

    if header is None:
        raise ValueError("缺少必要列：资产项, 金额(元)")

    # Find column indices
    name_col = next(i for i, h in enumerate(header) if h == "资产项")
    amount_col = next(
        (i for i, h in enumerate(header) if "金额" in h),
        None,
    )
    if amount_col is None:
        raise ValueError("缺少必要列：金额(元)")

    currency_col = next(
        (i for i, h in enumerate(header) if "币种" in h),
        None,
    )
    profit_col = next(
        (i for i, h in enumerate(header) if "收益" in h),
        None,
    )

    items: list[dict] = []
    for row in rows_data[header_idx + 1:]:
        if not row or all(cell is None for cell in row):
            continue

        name = str(row[name_col]).strip() if row[name_col] else ""
        if not name:
            continue

        try:
            raw_amount = row[amount_col]
            if isinstance(raw_amount, str):
                raw_amount = raw_amount.replace(",", "").replace("¥", "").replace("￥", "").strip()
            amount = float(raw_amount)
        except (ValueError, TypeError):
            logger.warning(f"Skipping row with invalid amount: {name}")
            continue

        if amount < 0:
            logger.warning(f"Skipping row with negative amount: {name} = {amount}")
            continue

        currency = "CNY"
        if currency_col is not None and row[currency_col]:
            currency = str(row[currency_col]).strip()

        item: dict = {
            "资产项": name,
            "金额(元)": amount,
            "币种": currency,
        }

        if profit_col is not None and row[profit_col] is not None:
            try:
                raw_profit = row[profit_col]
                if isinstance(raw_profit, str):
                    raw_profit = raw_profit.replace(",", "").replace("¥", "").replace("￥", "").strip()
                item["收益"] = float(raw_profit)
            except (ValueError, TypeError):
                pass

        items.append(item)

    return items


def import_from_parsed_data(
    db: Session,
    items: list[dict],
    file_name: str,
    record_date: date,
    source: str = "excel_upload",
    channel: str = "微众银行",
    weekly_investments: dict[str, float] | None = None,
    channel_weekly_total: float | None = None,
) -> ImportResult:
    """
    Import parsed asset data into the system. 有新数据直接覆盖旧数据。

    Args:
        db: Database session
        items: List of {"资产项": str, "金额(元)": float, "币种": str, "基金代码"?: str}
        file_name: Original file name for logging
        record_date: Portfolio record date (will be normalized to Monday)
        source: Import source type
        channel: Investment channel (e.g. "微众银行", "支付宝")
        weekly_investments: Optional {fund_code: amount} for weekly investment per fund
        channel_weekly_total: Optional total weekly investment for the entire channel
            (used when per-fund breakdown is unavailable, e.g. WeBank)

    Returns:
        ImportResult with import statistics
    """
    # Normalize record_date to Monday of the ISO week
    record_date = normalize_to_monday(record_date)

    total_items = len(items)
    matched_funds = 0
    new_funds_created = 0
    new_fund_ids: list[int] = []
    fund_records: list[tuple[Fund, float, str, float | None]] = []

    # Match or create funds
    for item in items:
        fund_name = item["资产项"]
        amount = item["金额(元)"]
        currency = item["币种"]
        profit_from_source = item.get("收益")
        fund_code = item.get("基金代码")

        fund, is_new = find_or_create_fund(db, fund_name, currency, fund_code=fund_code)

        if is_new:
            new_funds_created += 1
            new_fund_ids.append(fund.id)
        else:
            matched_funds += 1

        fund_records.append((fund, amount, currency, profit_from_source))

    # AI classification for new funds
    classification_results: dict = {"classified": 0, "models_covered": 0}
    if new_fund_ids:
        try:
            cls_map = classify_funds_with_ai(db, new_fund_ids)
            classification_results["classified"] = len(cls_map)
            if cls_map:
                all_models: set[int] = set()
                for fund_cls in cls_map.values():
                    all_models.update(fund_cls.keys())
                classification_results["models_covered"] = len(all_models)
        except Exception as e:
            logger.error(f"AI classification failed: {e}")

    # Get exchange rates for currency conversion
    usd_rate = get_latest_rate(db, "USD/CNY")
    hkd_rate = get_latest_rate(db, "HKD/CNY")

    # Write portfolio records (upsert)
    # Track pending records by fund_id to handle in-batch duplicates
    pending_records: dict[int, dict] = {}
    for fund, amount, currency, source_profit in fund_records:
        # Calculate amount_cny
        if currency == "USD":
            amount_cny = amount * usd_rate
        elif currency == "HKD":
            amount_cny = amount * hkd_rate
        else:
            amount_cny = amount

        # Only use profit when the source data explicitly provides it.
        profit = source_profit

        # Resolve weekly investment for this fund
        wi = None
        if weekly_investments and fund.code in weekly_investments:
            wi = weekly_investments[fund.code]

        # Use dict to deduplicate by fund_id (last write wins)
        pending_records[fund.id] = {
            "amount": amount,
            "amount_cny": amount_cny,
            "profit": profit,
            "weekly_investment": wi,
        }

    # Delete ALL existing records for the same (channel, week) then insert fresh.
    # This ensures data_completion leftovers or renamed funds are cleaned up.
    db.query(PortfolioRecord).filter(
        PortfolioRecord.channel == channel,
        PortfolioRecord.record_date == record_date,
    ).delete(synchronize_session=False)

    records_imported = 0
    for fund_id, data in pending_records.items():
        db.add(PortfolioRecord(
            fund_id=fund_id,
            record_date=record_date,
            channel=channel,
            amount=data["amount"],
            amount_cny=data["amount_cny"],
            profit=data["profit"],
            weekly_investment=data["weekly_investment"],
        ))
        records_imported += 1

    db.commit()

    # Generate snapshot
    snapshot_generated = False
    try:
        result = generate_snapshot(db, record_date)
        snapshot_generated = result is not None
    except Exception as e:
        logger.error(f"Snapshot generation failed: {e}")

    # Create import log
    # Compute total weekly investment: per-fund sum or channel-level total
    total_wi = channel_weekly_total
    if not total_wi and weekly_investments:
        total_wi = sum(weekly_investments.values())

    import_log = ImportLog(
        import_date=record_date,
        source=source,
        file_name=file_name,
        record_count=records_imported,
        new_funds_count=new_funds_created,
        status="success",
        weekly_investment_total=total_wi,
    )
    db.add(import_log)
    db.commit()
    db.refresh(import_log)

    return ImportResult(
        total_items=total_items,
        matched_funds=matched_funds,
        new_funds_created=new_funds_created,
        records_imported=records_imported,
        classification_results=classification_results,
        snapshot_generated=snapshot_generated,
        import_log_id=import_log.id,
    )


def import_from_excel(
    db: Session,
    file_content: bytes,
    file_name: str,
    record_date: date,
) -> ImportResult:
    """
    Full pipeline: parse Excel -> match/create funds -> AI classify -> write records -> snapshot.

    Args:
        db: Database session
        file_content: Excel file binary content
        file_name: Original file name
        record_date: Portfolio record date

    Returns:
        ImportResult with import statistics

    Raises:
        ValueError: File format error
    """
    items = _parse_excel(file_content)
    if not items:
        raise ValueError("Excel 文件中未找到有效的资产数据")

    return import_from_parsed_data(
        db=db,
        items=items,
        file_name=file_name,
        record_date=record_date,
        source="excel_upload",
    )
